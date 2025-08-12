"""
Gerador de Excel para Extract Fotos
Responsável por criar planilhas Excel organizadas com os dados extraídos
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Optional
from datetime import datetime
import os

# Importa as classes do parser
from parser import FileInfo, CondominioType

class ExcelGenerator:
    """Gerador de planilhas Excel para dados de condomínios"""
    
    def __init__(self):
        """Inicializa o gerador de Excel"""
        self.workbook = None
        self.worksheet = None
        
        # Cores para formatação
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_excel_from_files(self, files_info: List[FileInfo], output_filename: str = None) -> str:
        """
        Cria planilha Excel a partir dos dados dos arquivos
        
        Args:
            files_info: Lista de FileInfo processados
            output_filename: Nome do arquivo de saída (opcional)
            
        Returns:
            Nome do arquivo Excel gerado
        """
        if not output_filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"extract_fotos_{timestamp}.xlsx"
        
        # Cria o workbook e worksheet
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Dados Extraídos"
        
        # Determina o tipo de condomínio baseado nos dados
        condominio_type = self._determine_condominio_type(files_info)
        
        # Cria o DataFrame
        df = self._create_dataframe(files_info, condominio_type)
        
        # Adiciona dados ao Excel
        self._add_data_to_worksheet(df)
        
        # Aplica formatação
        self._apply_formatting(condominio_type)
        
        # Adiciona estatísticas
        self._add_statistics_sheet(files_info)
        
        # Salva o arquivo
        self.workbook.save(output_filename)
        print(f"✅ Planilha Excel criada: {output_filename}")
        
        return output_filename
    
    def _determine_condominio_type(self, files_info: List[FileInfo]) -> CondominioType:
        """Determina o tipo de condomínio baseado nos dados"""
        valid_files = [f for f in files_info if f.is_valid]
        
        if not valid_files:
            return CondominioType.COM_BLOCOS  # Default
        
        # Verifica se há arquivos com blocos
        has_blocks = any(f.condominio_type == CondominioType.COM_BLOCOS for f in valid_files)
        
        if has_blocks:
            return CondominioType.COM_BLOCOS
        else:
            return CondominioType.SEM_BLOCOS
    
    def _create_dataframe(self, files_info: List[FileInfo], condominio_type: CondominioType) -> pd.DataFrame:
        """Cria DataFrame pandas com os dados organizados"""
        data = []
        
        for file_info in files_info:
            if not file_info.is_valid:
                continue
                
            row = {
                'Nome do Arquivo': file_info.filename,
                'Apartamento': file_info.apartamento,
                'Leitura': file_info.leitura
            }
            
            # Adiciona coluna de bloco se necessário
            if condominio_type == CondominioType.COM_BLOCOS:
                row['Bloco'] = file_info.bloco if file_info.bloco else 'N/A'
            
            data.append(row)
        
        # Cria DataFrame
        df = pd.DataFrame(data)
        
        # Reorganiza colunas
        if condominio_type == CondominioType.COM_BLOCOS:
            df = df[['Nome do Arquivo', 'Bloco', 'Apartamento', 'Leitura']]
        else:
            df = df[['Nome do Arquivo', 'Apartamento', 'Leitura']]
        
        return df
    
    def _add_data_to_worksheet(self, df: pd.DataFrame):
        """Adiciona dados do DataFrame ao worksheet"""
        # Adiciona cabeçalho
        for col_num, column_title in enumerate(df.columns, 1):
            cell = self.worksheet.cell(row=1, column=col_num, value=column_title)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border
        
        # Adiciona dados
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _apply_formatting(self, condominio_type: CondominioType):
        """Aplica formatação ao worksheet"""
        # Ajusta largura das colunas
        column_widths = {
            'Nome do Arquivo': 30,
            'Bloco': 10,
            'Apartamento': 15,
            'Leitura': 15
        }
        
        for col_num, column_title in enumerate(self.worksheet[1], 1):
            if column_title.value in column_widths:
                self.worksheet.column_dimensions[column_title.column_letter].width = column_widths[column_title.value]
        
        # Remove coluna de bloco se não for necessária
        if condominio_type == CondominioType.SEM_BLOCOS:
            # Encontra a coluna do bloco e remove
            for col_num, column_title in enumerate(self.worksheet[1], 1):
                if column_title.value == 'Bloco':
                    self.worksheet.delete_cols(col_num)
                    break
    
    def _add_statistics_sheet(self, files_info: List[FileInfo]):
        """Adiciona planilha de estatísticas"""
        stats_ws = self.workbook.create_sheet("Estatísticas")
        
        # Calcula estatísticas
        total_files = len(files_info)
        valid_files = sum(1 for f in files_info if f.is_valid)
        invalid_files = total_files - valid_files
        
        # Estatísticas por tipo
        with_blocks = sum(1 for f in files_info if f.condominio_type == CondominioType.COM_BLOCOS and f.is_valid)
        without_blocks = sum(1 for f in files_info if f.condominio_type == CondominioType.SEM_BLOCOS and f.is_valid)
        
        # Dados para estatísticas
        stats_data = [
            ["Métrica", "Valor"],
            ["Total de Arquivos", total_files],
            ["Arquivos Válidos", valid_files],
            ["Arquivos Inválidos", invalid_files],
            ["COM Blocos", with_blocks],
            ["SEM Blocos", without_blocks],
            ["Taxa de Sucesso", f"{(valid_files/total_files*100):.1f}%" if total_files > 0 else "0%"],
            ["", ""],
            ["Data de Geração", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
            ["Arquivos com Erro", ""]
        ]
        
        # Adiciona arquivos com erro
        error_files = [f.filename for f in files_info if not f.is_valid]
        for error_file in error_files:
            stats_data.append(["", error_file])
        
        # Adiciona dados ao worksheet
        for row_num, row_data in enumerate(stats_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = stats_ws.cell(row=row_num, column=col_num, value=value)
                
                # Formata cabeçalho
                if row_num == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                
                cell.border = self.border
        
        # Ajusta largura das colunas
        stats_ws.column_dimensions['A'].width = 25
        stats_ws.column_dimensions['B'].width = 30

# Função de conveniência para uso direto
def generate_excel_report(files_info: List[FileInfo], output_filename: str = None) -> str:
    """
    Função simples para gerar relatório Excel
    
    Args:
        files_info: Lista de FileInfo processados
        output_filename: Nome do arquivo de saída (opcional)
        
    Returns:
        Nome do arquivo Excel gerado
    """
    generator = ExcelGenerator()
    return generator.create_excel_from_files(files_info, output_filename)

if __name__ == "__main__":
    # Teste básico da classe
    print("🧪 Testando Excel Generator...")
    
    # Cria dados de teste
    from parser import FileInfo, CondominioType
    
    test_files = [
        FileInfo("A-101-1234.jpg", CondominioType.COM_BLOCOS, "A", "101", "1234", True),
        FileInfo("B-205-5678.png", CondominioType.COM_BLOCOS, "B", "205", "5678", True),
        FileInfo("101-1234.jpg", CondominioType.SEM_BLOCOS, None, "101", "1234", True),
        FileInfo("205-5678.png", CondominioType.SEM_BLOCOS, None, "205", "5678", True),
        FileInfo("invalid-name.txt", CondominioType.COM_BLOCOS, None, None, None, False, "Padrão inválido")
    ]
    
    try:
        # Gera relatório
        output_file = generate_excel_report(test_files, "teste_extract_fotos.xlsx")
        print(f"✅ Relatório de teste gerado: {output_file}")
        
        # Verifica se o arquivo foi criado
        if os.path.exists(output_file):
            print(f"📁 Arquivo criado com sucesso: {os.path.getsize(output_file)} bytes")
        else:
            print("❌ Erro: Arquivo não foi criado")
            
    except Exception as e:
        print(f"❌ Erro durante teste: {e}")
